from __future__ import annotations
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from ..logger.adapters import LogAdapter

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from .models import SheetsLoadResult, ImportedSheetsRow
from base import ServiceBase


class SpreadsheetFileService(ServiceBase):

    def __init__(self, logger: LogAdapter):
        super().__init__(logger)

    def load(
        self, file_path: Path, required_headers: set | None = None
    ) -> SheetsLoadResult:

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            sheet = workbook.active

            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                message = "Sheet does not have any rows."
                self._logging(message, "ERROR")
                return SheetsLoadResult(
                    ok=False,
                    file_path=file_path,
                    message=message,
                )

            headers = [str(value).strip() if value else "" for value in rows[0]]
            imported_rows = []

            if required_headers:
                missing_headers = required_headers.difference(headers)
                if missing_headers:
                    message = f'Missing Required Headers: {", ".join(missing_headers)}'
                    self._logging(message, "ERROR")
                    return SheetsLoadResult(
                        ok=False,
                        file_path=file_path,
                        message=message,
                    )

            for excel_row_number, row_values in enumerate(rows[1:], start=2):
                row_data = {
                    headers[index]: value
                    for index, value in enumerate(row_values)
                    if index < len(headers)
                }
                ImportedSheetsRow(row_number=excel_row_number, values=row_data)
                imported_rows.append(
                    ImportedSheetsRow(row_number=excel_row_number, values=row_data)
                )
            message = f"File Successfully loaded {file_path}"
            self._logging(message, "INFO")
            return SheetsLoadResult(
                ok=True, file_path=file_path, rows=imported_rows, message=message
            )
        except FileNotFoundError:
            message = f"File not found: {file_path}"
            self._logging(message, "ERROR")
            return SheetsLoadResult(ok=False, file_path=file_path, message=message)

        except InvalidFileException as e:
            message = "Invalid Excel file"
            self._logging(message, "ERROR")
            self._logging(f"Invalid Excel file: {e}", "DEBUG")
            return SheetsLoadResult(ok=False, file_path=file_path, message=message)

        except PermissionError:
            message = f"Permission denied while reading file: {file_path}"
            self._logging(message, "ERROR")
            return SheetsLoadResult(ok=False, file_path=file_path, message=message)
        except Exception as e:
            message = "Unexpected error occurred import sheets file."
            self._logging(message, "ERROR")
            self._logging(f"Unexpected Error: {e}", "DEBUG")
            return SheetsLoadResult(ok=False, file_path=file_path, message=message)
